# pylint: disable=line-too-long, missing-module-docstring, import-error, too-many-arguments
import json
from openpyxl import Workbook
from data_collection import DataCollection


def load_dictionary_data(filename):
    """
    Load a dictionary from a JSON file.

    :param filename: The name of the JSON file to load.
    :return: The loaded dictionary.
    """
    with open(filename, encoding="utf-8") as f:
        return json.load(f)


def calculate_speed_stat_at_50(base, ivs, evs, nature, boosts):
    """
    Calculate a Pokémon's speed stat at level 50 given certain conditions.

    :param base: The base speed of the Pokémon.
    :param ivs: The Speed IVs of the Pokémon.
    :param evs: The Speed EVs of the Pokémon.
    :param nature: The nature of the Pokémon.
    :param boosts: The boosts to the Pokémon's speed.
    :return: The calculated speed stat.
    """
    # Correctly accounts for boosts
    boost_modifier = 1 + abs(boosts) / 2
    if boosts < 0:
        boost_modifier = boost_modifier ** -1

    return int(((2 * base + ivs + evs / 4) * 50 / 100 + 5) * nature * boost_modifier)


def add_speed_entry(name, base, ivs, evs, nature, boosts):
    """
    Adds a speed entry with the correct format to the all_data list. This is what is fully displayed on the sheet.

    :param name: The name of the Pokémon.
    :param base: The base speed of the Pokémon.
    :param ivs: The Speed IVs of the Pokémon.
    :param evs: The Speed EVs of the Pokémon.
    :param nature: The nature of the Pokémon.
    :param boosts: The boosts to the Pokémon's speed.
    """
    speed = calculate_speed_stat_at_50(base, ivs, evs, nature, boosts)

    # Converts Nature from a float multiplier to a string for a display
    for multiplier, effect in {0.9: "Negative", 1: "Neutral", 1.1: "Positive"}.items():
        if nature == multiplier:
            nature = effect

    if boosts > 0:
        boosts = "+" + str(boosts)

    master_data.append([speed, name, base, ivs, evs, nature, str(boosts)])

    if name in singles_pokemon:
        singles_data.append([speed, name, base, ivs, evs, nature, str(boosts)])


def construct_speed_tier_entry(name, base_speed):
    """
    Fully constructs all the speed tier entries for a Pokémon, given its base speed, and adds them to the all_data list.

    Configurations include the following:
    - absolute minimum speed (-spe nature and 0 IVs for trick room)
    - minimum dump speed (-spe nature but maxed IVs)
    - neutral dump speed (neutral speed nature and maxed IVs)
    - neutral speed invested (neutral speed nature maxed IVs and maxed EVs)
    - maximum speed (maxed evs ivs and positive speed nature)

    Every configuration is then repeated for every possible speed boost from -2 to +2.

    :param base_speed: The base speed of the Pokémon.
    :param name: The name of the Pokémon.
    """

    # Speed tiers, and accounting for boosts
    for boost in range(-2, 3):
        add_speed_entry(name, int(base_speed), 0, 0, 0.9, boost)
        add_speed_entry(name, int(base_speed), 31, 0, 0.9, boost)
        add_speed_entry(name, int(base_speed), 31, 0, 1, boost)
        add_speed_entry(name, int(base_speed), 31, 252, 1, boost)
        add_speed_entry(name, int(base_speed), 31, 252, 1.1, boost)


if __name__ == "__main__":
    print("This is a special script to establish the different speed tiers present in the game's metagame.")
    print("This will fill up one excel file: speedtiers.xlsx.\n")

    master_data = []
    singles_data = []

    print("Accessing the singles tier list...")
    singles_pokemon = []
    with open("singles_tierlist.txt", encoding="utf-8") as f:
        for line in f:
            if line.startswith("=") or line == "\n":
                continue

            line = line.strip()
            singles_pokemon.append(line)

    print("Constructing the speed tiers for normal Pokémon...")
    # A modified version of pokemon_info.json is used here, as we do not care about every Pokémon.
    pokemon = load_dictionary_data("pokemon_info.json")

    for key, value in pokemon.items():
        dc = DataCollection(value.split(",")[0], "../../gamedata/pokemon.txt")
        data = dc.extract_pokemon_data()

        base_spd = data["BaseStats"].split(",")[3]
        construct_speed_tier_entry(value.split(",")[1], base_spd)

    print("Constructing the speed tiers for alternate forms...")
    alt_forms = load_dictionary_data("alternate_forms.json")

    for key, value in alt_forms.items():
        construct_speed_tier_entry(key, value)

    print("Sorting the data...")
    master_data.sort(key=lambda x: (-x[0], -x[2], x[1]), reverse=False)
    singles_data.sort(key=lambda x: (-x[0], -x[2], x[1]), reverse=False)

    print("Grouping the data by base speed...")
    n = 1
    while n < len(singles_data):
        # If the Pokémon have the same speed stat and base speed, group them together
        if singles_data[n][0] == singles_data[n - 1][0] and singles_data[n][2] == singles_data[n - 1][2]:
            singles_data[n - 1][1] += ", " + singles_data[n][1]
            singles_data.remove(singles_data[n])
        else:
            n += 1

    print("Writing the data to an Excel file...")
    workbook = Workbook()
    singles_sheet = workbook.create_sheet("Singles")
    master_sheet = workbook.create_sheet("Master")
    workbook.remove(workbook["Sheet"])

    singles_sheet.append(["Speed", "Name", "Base", "IVs", "EVs", "Nature", "Boosts"])
    master_sheet.append(["Speed", "Name", "Base", "IVs", "EVs", "Nature", "Boosts"])

    for entry in singles_data:
        singles_sheet.append(entry)

    for entry in master_data:
        master_sheet.append(entry)

    workbook.save("speedtiers.xlsx")
